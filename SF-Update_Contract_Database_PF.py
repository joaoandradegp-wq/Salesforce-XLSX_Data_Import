#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import sys
import glob
import traceback
from datetime import datetime, date

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

# ==================== VARIÁVEIS GLOBAIS ====================

versao = "1.0.1"

# ===========================================================

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("A biblioteca 'openpyxl' nao esta instalada. Rode: pip install openpyxl")
    sys.exit(1)


# ----------------------------------------------------------------------
# Configuracoes fixas
# ----------------------------------------------------------------------

SOURCE_COLUMNS = [
    "SF Id Contrato",
    "Locavia Data de início do contrato",
    "Locavia Status Contrato",
    "Data Cancelamento",
]

TARGET_HEADERS = ["id", "StartDate", "Status", "IRIS_DataCancelamento__c"]

DATE_COLUMNS = {"StartDate", "IRIS_DataCancelamento__c"}

STATUS_MAP = {
    "Em Vigência": "Activated",
    "Aberto": "Draft",
    "Assinado": "Draft",
    "Em Assinatura": "Draft",
}

SHEET_NAME_PATTERN = re.compile(r"^\d{8}\s*-\s*LIVRE\s*-\s*Dados\s*Contra", re.IGNORECASE)

# Nome do arquivo de saida: DD-MM-AAAA.xlsx
OUTPUT_NAME_PATTERN = re.compile(r"^(\d{2})-(\d{2})-(\d{4})\.xlsx$", re.IGNORECASE)

LOG_FILENAME = "log_atualizacoes.txt"


def get_root_dir():
    """Retorna a pasta onde o programa esta rodando (raiz do programa)."""
    if getattr(sys, "frozen", False):
        # Executavel gerado (ex: PyInstaller)
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


# ----------------------------------------------------------------------
# Funcoes de transformacao
# ----------------------------------------------------------------------

def format_date(value):
    """Converte o valor para string no formato AAAA-MM-DD.
    'NULL' (em qualquer variacao de caixa) ou vazio vira string vazia."""
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()
    if text == "" or text.upper() == "NULL":
        return ""

    # 1) Extrai diretamente a parte da data, ignorando hora/milissegundos/
    #    qualquer sufixo que venha junto (ex: "2023-02-13 12:00:00.000").
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)  # AAAA-MM-DD...
    if m:
        y, mo, d = (int(g) for g in m.groups())
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            pass

    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})", text)  # DD/MM/AAAA...
    if m:
        d, mo, y = (int(g) for g in m.groups())
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            pass

    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})", text)  # DD-MM-AAAA...
    if m:
        d, mo, y = (int(g) for g in m.groups())
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            pass

    # 2) Fallback: tenta formatos completos conhecidos (com ou sem hora)
    known_formats = (
        "%d/%m/%Y %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S.%f",
        "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d",
    )
    for fmt in known_formats:
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    # Nao foi possivel interpretar como data - mantem o texto original
    return text


def map_status(value):
    """Aplica o de/para de status. Valores nao mapeados permanecem como estao."""
    if value is None:
        return ""
    text = str(value).strip()
    return STATUS_MAP.get(text, text)


def clean_id(value):
    if value is None:
        return ""
    return str(value).strip()


# ----------------------------------------------------------------------
# Leitura / processamento da planilha de origem
# ----------------------------------------------------------------------

def find_source_sheet(wb):
    for name in wb.sheetnames:
        if SHEET_NAME_PATTERN.match(name.strip()):
            return name
    return None


def read_and_process_source(ws, progress_callback=None):
    """Le a aba de origem inteira e retorna a lista de linhas processadas
    (lista de dicts na ordem TARGET_HEADERS).

    progress_callback(fracao), se informado, e chamado periodicamente com
    um valor entre 0.0 e 1.0 indicando o andamento da leitura."""

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("A aba de origem esta vazia.")

    header_row = rows[0]
    header_index = {}
    for idx, h in enumerate(header_row):
        if h is None:
            continue
        header_index[str(h).strip()] = idx

    missing = [c for c in SOURCE_COLUMNS if c not in header_index]
    if missing:
        raise ValueError(
            "As seguintes colunas nao foram encontradas na aba de origem:\n- "
            + "\n- ".join(missing)
        )

    col_positions = [header_index[c] for c in SOURCE_COLUMNS]

    data_rows = rows[1:]
    total = len(data_rows)
    step = max(total // 100, 1)  # atualiza a cada ~1% para nao travar a UI

    processed = []
    for i, data_row in enumerate(data_rows, start=1):
        # ignora linhas totalmente vazias
        if data_row is None or all(v is None for v in data_row):
            if progress_callback and (i % step == 0 or i == total):
                progress_callback(i / total if total else 1.0)
            continue

        raw_values = []
        for pos in col_positions:
            raw_values.append(data_row[pos] if pos < len(data_row) else None)

        sf_id, start_date, status, cancel_date = raw_values

        item = {
            "id": clean_id(sf_id),
            "StartDate": format_date(start_date),
            "Status": map_status(status),
            "IRIS_DataCancelamento__c": format_date(cancel_date),
        }
        processed.append(item)

        if progress_callback and (i % step == 0 or i == total):
            progress_callback(i / total if total else 1.0)

    return processed


# ----------------------------------------------------------------------
# Nome do arquivo de saida
# ----------------------------------------------------------------------

def extract_date_from_filename(filename):
    """Extrai AAAAMMDD do inicio do nome do arquivo original e retorna
    um objeto date. Ex: '20260817 - LIVRE - ...xlsx' -> date(2026, 8, 17)"""
    base = os.path.basename(filename)
    match = re.match(r"^(\d{8})", base)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y%m%d").date()
    except ValueError:
        return None


def build_output_filename(source_filename):
    dt = extract_date_from_filename(source_filename)
    if dt is None:
        # fallback: usa a data de hoje se nao conseguir extrair do nome
        dt = date.today()
    return dt.strftime("%d-%m-%Y") + ".xlsx", dt


# ----------------------------------------------------------------------
# Comparacao com execucao anterior
# ----------------------------------------------------------------------

def find_previous_output(data_dir):
    """Procura na pasta de dados (Contract/Data) por arquivos gerados em
    execucoes anteriores (padrao DD-MM-AAAA.xlsx) e retorna o caminho do
    mais recente, com base na data contida no proprio nome do arquivo.

    Nao ha exclusao por nome: se ja existir um arquivo com o mesmo nome do
    que sera gerado agora (reprocessamento do mesmo dia), ele deve mesmo
    assim ser considerado a execucao anterior para fins de comparacao."""
    candidates = []
    for path in glob.glob(os.path.join(data_dir, "*.xlsx")):
        name = os.path.basename(path)
        m = OUTPUT_NAME_PATTERN.match(name)
        if not m:
            continue
        try:
            d = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            continue
        candidates.append((d, path))

    if not candidates:
        return None

    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]


def load_previous_data(path):
    """Carrega a aba 'Salesforce' de um arquivo anterior e retorna um dict
    id -> {StartDate, Status, IRIS_DataCancelamento__c}."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return {}

    if "Salesforce" not in wb.sheetnames:
        return {}

    ws = wb["Salesforce"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    try:
        idx_id = header.index("id")
        idx_start = header.index("StartDate")
        idx_status = header.index("Status")
        idx_cancel = header.index("IRIS_DataCancelamento__c")
    except ValueError:
        return {}

    data = {}
    for row in rows[1:]:
        if row is None or all(v is None for v in row):
            continue
        rid = clean_id(row[idx_id] if idx_id < len(row) else None)
        if rid == "":
            continue
        data[rid] = {
            "StartDate": row[idx_start] if idx_start < len(row) else "",
            "Status": row[idx_status] if idx_status < len(row) else "",
            "IRIS_DataCancelamento__c": row[idx_cancel] if idx_cancel < len(row) else "",
        }
    return data


def compare_data(current_rows, previous_data):
    """Compara os registros atuais com os da execucao anterior.
    Retorna (total, novos, atualizados, changed_rows), onde changed_rows
    e a lista (na mesma ordem/estrutura de current_rows) contendo apenas
    os registros novos ou que sofreram alguma alteracao - e essa lista
    que deve ser usada para importacao."""
    total = len(current_rows)
    novos = 0
    atualizados = 0
    changed_rows = []

    for item in current_rows:
        rid = item["id"]
        if rid == "":
            continue
        if rid not in previous_data:
            novos += 1
            changed_rows.append(item)
            continue

        prev = previous_data[rid]
        changed = (
            str(prev.get("StartDate") or "") != str(item["StartDate"] or "")
            or str(prev.get("Status") or "") != str(item["Status"] or "")
            or str(prev.get("IRIS_DataCancelamento__c") or "") != str(item["IRIS_DataCancelamento__c"] or "")
        )
        if changed:
            atualizados += 1
            changed_rows.append(item)

    return total, novos, atualizados, changed_rows


def write_log(log_dir, ref_date, total, novos, atualizados):
    log_path = os.path.join(log_dir, LOG_FILENAME)
    now = datetime.now()
    linha = (
        f"[{now.strftime('%d/%m/%Y %H:%M:%S')}] "
        f"Dia {ref_date.strftime('%d/%m/%Y')}, com {total} registros: "
        f"{novos} novos registros e {atualizados} registros atualizados.\n"
    )
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(linha)
    return linha


# ----------------------------------------------------------------------
# Escrita da nova aba / salvamento
# ----------------------------------------------------------------------

def write_salesforce_sheet(wb, processed_rows, progress_callback=None):
    if "Salesforce" in wb.sheetnames:
        del wb["Salesforce"]
    ws = wb.create_sheet("Salesforce")

    ws.append(TARGET_HEADERS)

    total = len(processed_rows)
    step = max(total // 100, 1)
    for i, item in enumerate(processed_rows, start=1):
        ws.append([item[h] for h in TARGET_HEADERS])
        if progress_callback and (i % step == 0 or i == total):
            progress_callback(i / total if total else 1.0)

    # ajusta largura das colunas de forma simples
    for i, header in enumerate(TARGET_HEADERS, start=1):
        col_letter = get_column_letter(i)
        max_len = max(
            [len(header)] + [len(str(item[header])) for item in processed_rows]
        ) if processed_rows else len(header)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    return ws


def copy_rows_to_clipboard(root, processed_rows):
    """Copia os dados (com cabecalho) para a area de transferencia, em formato
    separado por TAB - igual ao que o Excel gera ao selecionar celulas e dar CTRL+C."""
    lines = ["\t".join(TARGET_HEADERS)]
    for item in processed_rows:
        lines.append("\t".join(str(item[h]) for h in TARGET_HEADERS))
    text = "\n".join(lines)

    root.clipboard_clear()
    root.clipboard_append(text)
    root.update()  # garante que o conteudo fique disponivel na area de transferencia


# ----------------------------------------------------------------------
# Interface grafica
# ----------------------------------------------------------------------

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Update Contract Database "+versao+" - Aggrandize - João Márcio Bicalho Andrade")
        self.root.geometry("720x480")
        self.root.resizable(True, True)

        title = tk.Label(
            root,
            text="Atualizador de Base Salesforce",
            font=("Segoe UI", 14, "bold"),
        )
        title.pack(pady=(15, 5))

        subtitle = tk.Label(
            root,
            text="Selecione a planilha XLSX de origem para processar.",
            font=("Segoe UI", 10),
        )
        subtitle.pack(pady=(0, 10))

        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=5)

        self.btn_load = tk.Button(
            btn_frame,
            text="Selecionar Planilha XLSX",
            command=self.on_load_clicked,
            font=("Segoe UI", 10, "bold"),
            padx=12,
            pady=6,
        )
        self.btn_load.pack()

        progress_frame = tk.Frame(root)
        progress_frame.pack(fill=tk.X, padx=15, pady=(10, 0))

        self.progress_var = tk.DoubleVar(value=0.0)
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            variable=self.progress_var,
        )
        self.progress_bar.pack(fill=tk.X, side=tk.LEFT, expand=True)

        self.progress_label = tk.Label(progress_frame, text="0%", width=6, font=("Segoe UI", 9))
        self.progress_label.pack(side=tk.LEFT, padx=(8, 0))

        self.log_area = scrolledtext.ScrolledText(
            root, wrap=tk.WORD, font=("Consolas", 10), height=20
        )
        self.log_area.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        self.log_area.configure(state=tk.DISABLED)

        self.root_dir = get_root_dir()
        self.log_dir = os.path.join(self.root_dir, "Contract")
        self.data_dir = os.path.join(self.log_dir, "Data")
        os.makedirs(self.data_dir, exist_ok=True)

        self.log(f"Pasta raiz do programa: {self.root_dir}")
        self.log(f"Planilhas serao salvas em: {self.data_dir}")
        self.log(f"Log sera salvo em: {self.log_dir}")

    def log(self, message):
        self.log_area.configure(state=tk.NORMAL)
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.log_area.configure(state=tk.DISABLED)
        self.root.update_idletasks()

    def set_progress(self, value):
        """Atualiza a barra de progresso. value vai de 0 a 100."""
        value = max(0.0, min(100.0, value))
        self.progress_var.set(value)
        self.progress_label.configure(text=f"{int(value)}%")
        self.root.update_idletasks()

    def make_stage_callback(self, start, end):
        """Retorna uma funcao que recebe uma fracao (0.0-1.0) e mapeia
        para o intervalo [start, end] da barra de progresso."""
        def callback(fraction):
            self.set_progress(start + (end - start) * fraction)
        return callback

    def on_load_clicked(self):
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha XLSX",
            filetypes=[("Arquivos Excel", "*.xlsx")],
        )
        if not file_path:
            return

        self.btn_load.configure(state=tk.DISABLED)
        self.set_progress(0)
        try:
            self.process_file(file_path)
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("Erro", str(exc))
            self.log(f"ERRO: {exc}")
            self.set_progress(0)
        finally:
            self.btn_load.configure(state=tk.NORMAL)

    def process_file(self, file_path):
        self.log("")
        self.log(f"Arquivo selecionado: {os.path.basename(file_path)}")
        self.set_progress(0)

        # --- Carregar workbook (0-10%) ---
        wb = openpyxl.load_workbook(file_path, data_only=True)
        self.set_progress(10)

        sheet_name = find_source_sheet(wb)
        if sheet_name is None:
            raise ValueError(
                "Nao foi encontrada nenhuma aba no formato "
                "'AAAAMMDD - LIVRE - Dados Contra...' dentro da planilha."
            )
        self.log(f"Aba de origem identificada: {sheet_name}")
        self.set_progress(15)

        # --- Leitura e processamento das linhas (15-55%) ---
        ws = wb[sheet_name]
        processed_rows = read_and_process_source(
            ws, progress_callback=self.make_stage_callback(15, 55)
        )
        self.set_progress(55)
        self.log(f"Registros lidos e tratados: {len(processed_rows)}")

        output_filename, ref_date = build_output_filename(os.path.basename(file_path))
        output_path = os.path.join(self.data_dir, output_filename)

        # --- Comparacao com execucao anterior (55-65%) e gravacao do log (sempre) ---
        previous_path = find_previous_output(self.data_dir)
        if previous_path:
            self.log(f"Planilha de execucao anterior encontrada: {os.path.basename(previous_path)}")
            previous_data = load_previous_data(previous_path)
            total, novos, atualizados, changed_rows = compare_data(processed_rows, previous_data)
            self.log("Comparacao concluida:")
        else:
            self.log("Nenhuma execucao anterior encontrada em Contract\\Data (primeira execucao).")
            total = len(processed_rows)
            novos = total
            atualizados = 0
            changed_rows = processed_rows

        log_line = write_log(self.log_dir, ref_date, total, novos, atualizados)
        self.log(f"  - Total de registros: {total}")
        self.log(f"  - Novos registros: {novos}")
        self.log(f"  - Registros atualizados: {atualizados}")
        self.log(f"Log gravado em: {LOG_FILENAME}")
        self.log(log_line.strip())
        self.set_progress(65)

        # --- Criacao da aba Salesforce (65-85%) ---
        write_salesforce_sheet(
            wb, processed_rows, progress_callback=self.make_stage_callback(65, 85)
        )
        self.set_progress(85)

        # --- Salvamento (85-95%) ---
        wb.save(output_path)
        self.log(f"Planilha salva em: {output_path}")
        self.set_progress(95)

        # --- Copia para a area de transferencia (95-100%) ---
        # Apenas os registros novos/atualizados devem ser importados
        copy_rows_to_clipboard(self.root, changed_rows)
        self.log(f"Registros copiados para a area de transferencia (novos + atualizados): {len(changed_rows)}")
        self.set_progress(100)

        if changed_rows:
            messagebox.showinfo(
                "Concluido",
                f"{len(changed_rows)} registro(s) novo(s)/atualizado(s) copiado(s) para a "
                "área de transferência (CTRL+C) para ser colado (CTRL+V) via DataImport",
            )
        else:
            messagebox.showinfo(
                "Concluido",
                "Nenhum registro novo ou atualizado desde a ultima execucao. "
                "Nada foi copiado para a area de transferencia.",
            )


def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()


if __name__ == "__main__":
    main()